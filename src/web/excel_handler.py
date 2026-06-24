"""Excel file parsing and generation for the search system.

Input Excel columns: 型号, 品牌, 数量(采购数量)
Output Excel columns: 型号, 品牌, 采购数量, 适用价格(人民币), 库存数量,
                      货期, 来源网站, 渠道链接, 原始币种价格, 查询时间, 最低价
"""

from __future__ import annotations

import logging
from pathlib import Path

from src.scheduler import SearchItem, SearchResultRow

logger = logging.getLogger(__name__)


def parse_upload_excel(file_path: Path) -> list[SearchItem]:
    """Parse the uploaded Excel file into SearchItem list.

    Expected columns (flexible matching):
    - 型号 / MPN / Part Number / 元器件型号
    - 品牌 / Brand / Manufacturer (optional)
    - 数量 / 采购数量 / Quantity / QTY (optional, default=1)
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("Excel文件没有活动工作表")

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip().lower())

    # Map columns
    mpn_col = _find_column(headers, ["型号", "mpn", "part number", "元器件型号", "partnumber", "part_number"])
    brand_col = _find_column(headers, ["品牌", "brand", "manufacturer", "厂商", "厂家"])
    qty_col = _find_column(headers, ["数量", "采购数量", "quantity", "qty", "购买数量"])

    if mpn_col is None:
        raise ValueError("找不到'型号'列（支持: 型号/MPN/Part Number/元器件型号）")

    items: list[SearchItem] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx > 10000:
            break

        mpn_val = row[mpn_col] if mpn_col < len(row) else None
        if not mpn_val or not str(mpn_val).strip():
            continue

        brand_val = ""
        if brand_col is not None and brand_col < len(row):
            brand_val = str(row[brand_col] or "").strip()

        qty_val = 1
        if qty_col is not None and qty_col < len(row):
            try:
                qty_val = int(float(str(row[qty_col] or "1")))
            except (ValueError, TypeError):
                qty_val = 1

        items.append(SearchItem(
            mpn=str(mpn_val).strip(),
            brand=brand_val,
            quantity=max(qty_val, 1),
            row_index=row_idx,
        ))

    wb.close()
    return items


def generate_result_excel(results: list[SearchResultRow], output_path: Path) -> None:
    """Generate the output Excel with all search results.

    Columns: 型号, 品牌, 采购数量, 适用价格(人民币), 库存数量,
             货期, 来源网站, 渠道链接, 原始币种价格, 查询时间, 最低价
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "比价结果"

    # Header
    headers = [
        "型号", "品牌", "采购数量", "适用价格(人民币)", "库存数量",
        "货期", "来源网站", "渠道链接", "原始币种价格", "查询时间", "最低价", "状态"
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    best_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result.mpn)
        ws.cell(row=row_idx, column=2, value=result.brand)
        ws.cell(row=row_idx, column=3, value=result.quantity)
        ws.cell(row=row_idx, column=4, value=result.price_cny)
        ws.cell(row=row_idx, column=5, value=result.stock)
        ws.cell(row=row_idx, column=6, value=result.lead_time or "未显示")
        ws.cell(row=row_idx, column=7, value=result.supplier)
        ws.cell(row=row_idx, column=8, value=result.product_url)
        ws.cell(row=row_idx, column=9, value=result.price_original)
        ws.cell(row=row_idx, column=10, value=result.query_time)
        ws.cell(row=row_idx, column=11, value="⭐" if result.is_best_price else "")
        ws.cell(row=row_idx, column=12, value=result.status)

        # Highlight best price row
        if result.is_best_price:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = best_fill

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for row in range(2, min(len(results) + 2, 50)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_len + 2

    wb.save(output_path)
    logger.info(f"Result Excel saved: {output_path} ({len(results)} rows)")


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    """Find column index matching any of the candidate names."""
    for i, header in enumerate(headers):
        for candidate in candidates:
            if candidate in header or header in candidate:
                return i
    return None
