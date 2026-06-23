import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging
from datetime import datetime
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

def read_purchase_list(file_path_or_stream) -> List[Dict[str, Any]]:
    """
    Reads the Excel file and extracts parts info.
    Returns a list of dicts: [{'model': str, 'brand': str, 'quantity': int}]
    """
    try:
        df = pd.read_excel(file_path_or_stream)
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise ValueError(f"无法读取Excel文件: {e}")

    # Standardize column names
    col_mapping = {}
    for col in df.columns:
        col_str = str(col).strip()
        if col_str in ["型号", "Part Number", "Part Number/型号", "Material", "物料型号"]:
            col_mapping[col] = "model"
        elif col_str in ["品牌", "Brand", "品牌/Manufacturer", "厂牌", "制造商", "Manufacturer", "厂商", "生产商"]:
            col_mapping[col] = "brand"
        elif col_str in ["数量", "采购数量", "Quantity", "Qty", "QTY", "需求数量", "总数量", "采购数", "总需数量", "数量(pcs)", "用量"]:
            col_mapping[col] = "quantity"

    # Validate mandatory columns
    required = ["model", "quantity"]
    mapped_cols = {v: k for k, v in col_mapping.items()}
    
    missing = [r for r in required if r not in mapped_cols]
    if missing:
        raise ValueError(f"Excel文件中缺少必要列，请确保包含以下列名之一: \n"
                         f"型号: {['型号', 'Part Number']}\n"
                         f"采购数量: {['数量', '采购数量', 'Quantity']}")

    df = df.rename(columns=col_mapping)
    
    # Fill missing brands with empty string
    if "brand" not in df.columns:
        df["brand"] = ""
    else:
        df["brand"] = df["brand"].fillna("").astype(str).str.strip()

    # Clean data
    df["model"] = df["model"].fillna("").astype(str).str.strip()
    df = df[df["model"] != ""]  # Filter out empty model rows
    
    # Parse quantity to integer
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(1).astype(int)

    result = df[["model", "brand", "quantity"]].to_dict(orient="records")
    logger.info(f"Successfully loaded {len(result)} items from Excel.")
    return result

def save_comparison_results(results: List[Dict[str, Any]], output_path: str):
    """
    Saves the list of dictionaries to an styled Excel spreadsheet.
    Identifies the lowest price for each model and highlights it.
    """
    if not results:
        logger.warning("No results to save.")
        # Create empty excel
        df = pd.DataFrame(columns=[
            "型号", "品牌", "采购数量", "适用价格(人民币)", 
            "库存数量", "货期", "来源网站", "渠道链接", "原始币种价格", "查询时间", "最低价"
        ])
        df.to_excel(output_path, index=False)
        return

    df = pd.DataFrame(results)

    # Standardize columns
    expected_cols = [
        "型号", "品牌", "采购数量", "适用价格(人民币)", 
        "库存数量", "货期", "来源网站", "渠道链接", "原始币种价格", "查询时间"
    ]
    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""

    # Sort results
    df = df[expected_cols]
    df = df.sort_values(by=["型号", "适用价格(人民币)"], ascending=[True, True])

    # Mark the lowest price for each model
    df["最低价"] = ""
    
    # Find the index of the minimum non-empty price for each model
    # Convert '适用价格(人民币)' to numeric for comparison
    df["_price_num"] = pd.to_numeric(df["适用价格(人民币)"], errors="coerce")
    
    # Group by model and find the index of the minimum price
    # Exclude NaN values (failed queries)
    for model, group in df.groupby("型号"):
        valid_prices = group.dropna(subset=["_price_num"])
        if not valid_prices.empty:
            min_idx = valid_prices["_price_num"].idxmin()
            df.loc[min_idx, "最低价"] = "⭐"

    # Drop temporary column
    df = df.drop(columns=["_price_num"])

    # Write to Excel
    df.to_excel(output_path, index=False)

    # Apply styling using openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Style templates
    font_family = "Microsoft YaHei"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    lowest_price_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light green
    lowest_price_font = Font(name=font_family, size=10, bold=True, color="375623")
    
    normal_font = Font(name=font_family, size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Style Header Row
    ws.row_dimensions[1].height = 26
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Style Data Rows
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        is_lowest = ws.cell(row=row_idx, column=ws.max_column).value == "⭐"
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 2]:  # Model, Brand
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [3, 4]:  # Quantity, Price
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Format price as currency if numeric
                if col_idx == 4 and isinstance(cell.value, (int, float)):
                    cell.number_format = '¥#,##0.00'
            elif col_idx in [5, 6, 7, 10, 11]:  # Stock, Lead Time, Site, Time, Lowest Price Indicator
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Apply Lowest Price styling or normal styling
            if is_lowest:
                cell.fill = lowest_price_fill
                if col_idx == 4 or col_idx == 11:
                    cell.font = lowest_price_font
                else:
                    cell.font = normal_font
            else:
                cell.font = normal_font

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # Give links a reasonable width
            if val_str.startswith("http"):
                max_len = max(max_len, 15)
            else:
                # Handle Chinese characters (count double length)
                str_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                max_len = max(max_len, str_len)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save styled workbook
    wb.save(output_path)
    logger.info(f"Excel comparison report successfully saved and styled at: {output_path}")

if __name__ == "__main__":
    # Test stub
    test_results = [
        {"型号": "STM32F103C8T6", "品牌": "ST", "采购数量": 100, "适用价格(人民币)": 12.5, "库存数量": 5000, "货期": "现货", "来源网站": "立创商城", "渠道链接": "http://szlcsc.com", "原始币种价格": "12.5 CNY", "查询时间": "2026-06-23"},
        {"型号": "STM32F103C8T6", "品牌": "ST", "采购数量": 100, "适用价格(人民币)": 11.2, "库存数量": 1000, "货期": "3天", "来源网站": "云汉芯城", "渠道链接": "http://ickey.cn", "原始币种价格": "11.2 CNY", "查询时间": "2026-06-23"},
        {"型号": "MAX232", "品牌": "TI", "采购数量": 10, "适用价格(人民币)": 3.5, "库存数量": 50, "货期": "现货", "来源网站": "立创商城", "渠道链接": "http://szlcsc.com", "原始币种价格": "3.5 CNY", "查询时间": "2026-06-23"}
    ]
    save_comparison_results(test_results, "test_output.xlsx")
    print("Done")
